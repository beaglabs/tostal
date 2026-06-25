import os
import shutil
import zipfile
import hashlib
import urllib.request
from pathlib import Path
from typing import Optional
import warnings

import numpy as np
import torch

try:
    import lasio
except ImportError:
    lasio = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

FORCE2020_URL = "https://zenodo.org/records/4351156/files/LAS_files_Force_2020_all_wells_train_test_blind_hidden_final.zip?download=1"
LITHOLOGY_URL = "https://zenodo.org/records/4351156/files/lithology%20scoring%20matrix%20cost%20function.xlsx?download=1"
NPD_LITHO_URL = "https://zenodo.org/records/4351156/files/NPD_Lithostratigraphy_groups_all_wells.xlsx?download=1"

CURVE_ALIASES = {
    "GR": ["GR", "GAM", "GR_R3", "GRC", "SGR", "CGR", "GRD", "GAMMA"],
    "RT": ["RT", "RD", "RDEP", "RLA5", "AT90", "ILD", "LLD", "RESD", "RLLD"],
    "RHOB": ["RHOB", "RHOZ", "DEN", "DENS", "RHO", "RHOZ", "ZDEN"],
    "NPHI": ["NPHI", "NPH", "NEU", "NEUT", "TNPH", "NPOR", "PHIN"],
    "DT": ["DT", "DTC", "AC", "DTP", "DT4P", "DTCO", "ACOUSTIC"],
    "CALI": ["CALI", "CAL", "CALIPER", "HD", "BS", "BIT"],
}

FACIES_CLASSES = [
    "Sandstone", "Sandstone/Shale", "Shale", "Marl", "Dolomite",
    "Limestone", "Chalk", "Halite", "Anhydrite", "Tuff",
    "Coal", "Basement",
]

FORCE_LITHOLOGY_CODE_TO_IDX = {
    30000: 0,   # Sandstone
    65030: 1,   # Sandstone/Shale
    65000: 2,   # Shale
    80000: 3,   # Marl
    74000: 4,   # Dolomite
    70000: 5,   # Limestone
    70032: 6,   # Chalk
    88000: 7,   # Halite
    86000: 8,   # Anhydrite
    99000: 9,   # Tuff
    90000: 10,  # Coal
    93000: 11,  # Basement
}


def _download(url, dest, desc="downloading"):
    os.makedirs(os.path.dirname(dest) or ".", exist_ok=True)
    if os.path.exists(dest):
        print(f"  {desc}: already cached at {dest}")
        return
    print(f"  {desc}: {url[:80]}...")
    urllib.request.urlretrieve(url, dest)
    print(f"  {desc}: saved to {dest}")


def _match_curve(las_file, target):
    candidates = CURVE_ALIASES[target]
    for c in candidates:
        if c in las_file.keys():
            return las_file[c].data
    return None


def download_force2020(cache_dir="data/force2020"):
    cache = Path(cache_dir)
    cache.mkdir(parents=True, exist_ok=True)
    las_zip = cache / "force2020_las.zip"
    litho_xlsx = cache / "lithology_scoring.xlsx"
    _download(FORCE2020_URL, las_zip, "FORCE 2020 LAS")
    _download(LITHOLOGY_URL, litho_xlsx, "lithology labels")
    return cache


def parse_las_files(las_dir, n_depth=512, n_curves=6):
    if lasio is None:
        raise ImportError("lasio required: pip install lasio")
    las_dir = Path(las_dir)
    las_paths = sorted(las_dir.rglob("*.las")) + sorted(las_dir.rglob("*.LAS"))
    if not las_paths:
        raise FileNotFoundError(f"No LAS files found recursively in {las_dir}")

    wells = {}
    for las_path in las_paths:
        well_name = las_path.stem
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            las = lasio.read(las_path)

        curve_data = np.zeros((n_curves, n_depth), dtype=np.float32)
        depths = None
        for i, key in enumerate(["GR", "RT", "RHOB", "NPHI", "DT", "CALI"]):
            data = _match_curve(las, key)
            if data is not None:
                if depths is None:
                    if las.index is not None:
                        depths = las.index
                    else:
                        depths = data  # fallback, use data index

        if depths is None:
            continue

        for i, key in enumerate(["GR", "RT", "RHOB", "NPHI", "DT", "CALI"]):
            data = _match_curve(las, key)
            if data is not None:
                data = np.asarray(data, dtype=np.float32)
            if data is not None and len(data) == len(depths):
                valid = ~np.isnan(data)
                if valid.sum() > 10:
                    data[~valid] = np.nanmean(data[valid]) if valid.any() else 0
                    f = np.linspace(0, len(data) - 1, n_depth)
                    curve_data[i] = np.interp(f, np.arange(len(data)), data)

        valid_mask = curve_data.sum(axis=0) != 0
        if valid_mask.sum() < n_depth * 0.1:
            continue

        mean = curve_data.mean(axis=1, keepdims=True) + 1e-8
        std = curve_data.std(axis=1, keepdims=True) + 1e-8
        curve_data = (curve_data - mean) / std
        curve_data[:, ~valid_mask] = 0

        facies = _read_lithology_from_las(las, depths, n_depth)

        wells[well_name] = {
            "well_log": torch.from_numpy(curve_data),
            "metadata": {
                "name": well_name,
                "basin": "Norwegian Sea",
            },
        }
        if facies is not None:
            wells[well_name]["facies"] = torch.from_numpy(facies)
    return wells


def _read_lithology_from_las(las, depths, n_depth):
    lith_candidates = ["FORCE_2020_LITHOFACIES_LITHOLOGY", "LITH", "LITHOLOGY", "FACIES"]
    for name in lith_candidates:
        if name in las.keys():
            raw = np.asarray(las[name].data, dtype=np.float32)
            class_idx = np.full(len(raw), -1, dtype=np.int64)
            for code, idx in FORCE_LITHOLOGY_CODE_TO_IDX.items():
                class_idx[np.isclose(raw, float(code), atol=1.0)] = idx
            src_x = np.arange(len(class_idx))
            target_x = np.linspace(0, len(class_idx) - 1, n_depth)
            labels = np.full(n_depth, -1, dtype=np.int64)
            for j in range(n_depth):
                nearest = int(round(target_x[j]))
                nearest = min(max(nearest, 0), len(class_idx) - 1)
                labels[j] = class_idx[nearest]
            return labels
    return None


def load_lithology_labels(xlsx_path, wells, n_depth=512):
    if openpyxl is None:
        raise ImportError("openpyxl required: pip install openpyxl")
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    name_col = next((i for i, h in enumerate(header) if "well" in h), None)
    top_col = next((i for i, h in enumerate(header) if "top" in h or "depth" in h), None)
    base_col = next((i for i, h in enumerate(header) if "base" in h or "bottom" in h), None)
    lith_col = next((i for i, h in enumerate(header) if "lith" in h or "facies" in h), None)

    if None in (name_col, top_col, base_col, lith_col):
        print(f"  Warning: could not identify columns in lithology xlsx. Header: {header}")
        return

    litho_label_to_idx = {name.lower(): i for i, name in enumerate(FACIES_CLASSES)}

    for row in rows[1:]:
        if not row or not row[name_col]:
            continue
        well_name = str(row[name_col]).strip()
        try:
            top = float(row[top_col])
            base = float(row[base_col])
            lith = str(row[lith_col]).strip().lower()
        except (ValueError, TypeError):
            continue

        if well_name not in wells or lith not in litho_label_to_idx:
            continue

        label = litho_label_to_idx[lith]
        well = wells[well_name]
        if "facies" not in well:
            well["facies"] = torch.full((n_depth,), -1, dtype=torch.long)

        depth_range = base - top
        if depth_range <= 0:
            continue
        n_points = int(depth_range / n_depth * 100) if depth_range < n_depth else n_depth
        top_idx = max(0, min(n_depth - 1, int(top / depth_range * n_depth)))
        base_idx = max(top_idx + 1, min(n_depth, int(base / depth_range * n_depth)))
        well["facies"][top_idx:base_idx] = label


def load_lithostratigraphy(xlsx_path, wells, n_depth=512):
    if openpyxl is None:
        return
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    name_col = next((i for i, h in enumerate(header) if "well" in h), None)
    group_col = next((i for i, h in enumerate(header) if "group" in h), None)
    top_col = next((i for i, h in enumerate(header) if "top" in h), None)
    base_col = next((i for i, h in enumerate(header) if "base" in h), None)

    if None in (name_col, group_col):
        return

    for row in rows[1:]:
        if not row or not row[name_col]:
            continue
        well_name = str(row[name_col]).strip()
        if well_name not in wells:
            continue

        group = str(row[group_col]).strip() if row[group_col] else ""
        try:
            top = float(row[top_col]) if top_col is not None and row[top_col] else 0
            base = float(row[base_col]) if base_col is not None and row[base_col] else n_depth
        except (ValueError, TypeError):
            continue

        well = wells[well_name]
        well.setdefault("formations", []).append({
            "name": group, "top": top, "base": base,
        })


def generate_geology_text_from_formations(well, depth_scale=1.0):
    formations = well.get("formations", [])
    if not formations:
        return "No formation data available."

    name = well["metadata"]["name"]
    n_units = len(formations)
    descs = []
    for f in formations:
        fname = f["name"]
        top = int(f["top"] * depth_scale)
        base = int(f["base"] * depth_scale)
        descs.append(f"{fname} from {top} to {base}m")

    unit_descriptions = ". ".join(descs) + "."
    text = f"Well {name} contains {n_units} lithostratigraphic units: {unit_descriptions}"
    return text


def process_force2020(
    cache_dir="data/force2020",
    n_depth=512,
    n_curves=6,
    force_reprocess=False,
):
    cache = Path(cache_dir)
    cache.mkdir(parents=True, exist_ok=True)
    processed_path = cache / "processed.pt"

    if processed_path.exists() and not force_reprocess:
        print(f"Loading cached processed data from {processed_path}")
        return torch.load(processed_path, weights_only=False)

    cache = download_force2020(cache_dir)
    las_zip = cache / "force2020_las.zip"
    las_dir = cache / "las"
    las_files = list(las_dir.rglob("*.las")) + list(las_dir.rglob("*.LAS")) if las_dir.exists() else []
    if not las_files:
        if las_dir.exists():
            shutil.rmtree(las_dir)
        print("Extracting LAS files...")
        las_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(las_zip, "r") as zf:
            zf.extractall(las_dir)

        # Flatten wrapper directory if zip has a top-level folder
        children = list(las_dir.iterdir())
        wrapper = None
        for child in children:
            if child.is_dir() and (child / "train").exists():
                wrapper = child
                break
        if wrapper:
            print(f"  Flattening wrapper directory {wrapper.name}")
            for item in wrapper.iterdir():
                target = las_dir / item.name
                if not target.exists():
                    shutil.move(str(item), str(target))
            shutil.rmtree(wrapper)

    train_las_dir = las_dir / "train"
    if not train_las_dir.exists():
        las_dirs = [p.parent for p in sorted(las_dir.rglob("*.las")) + sorted(las_dir.rglob("*.LAS"))]
        if las_dirs and las_dirs[0].exists():
            train_las_dir = las_dirs[0]
        else:
            train_las_dir = las_dir

    print(f"Parsing LAS files from {train_las_dir}...")
    wells = parse_las_files(train_las_dir, n_depth=n_depth, n_curves=n_curves)
    print(f"  Parsed {len(wells)} wells")

    litho_path = cache / "lithology_scoring.xlsx"
    if not litho_path.exists():
        _download(LITHOLOGY_URL, litho_path, "lithology labels")

    labeled = sum(1 for w in wells.values() if "facies" in w)
    print(f"  Loaded lithology labels from LAS files for {labeled} wells")

    npd_litho = cache / "npd_lithostratigraphy.xlsx"
    if not npd_litho.exists():
        _download(NPD_LITHO_URL, npd_litho, "NPD lithostratigraphy")
    if npd_litho.exists():
        print("Loading lithostratigraphy...")
        load_lithostratigraphy(npd_litho, wells, n_depth=n_depth)
        strat = sum(1 for w in wells.values() if "formations" in w)
        print(f"  Added stratigraphy to {strat} wells")

    torch.save(wells, processed_path)
    print(f"Saved {len(wells)} processed wells to {processed_path}")
    return wells


def build_spatial_features(curves, n_depth, window=5):
    """Build spatial features from well log curves for facies classification.

    Computes per-depth-point features that capture geological context:
    - Original 6 curve values
    - Normalized depth position
    - Rolling-window mean per curve (captures formation-scale trends)
    - Rolling-window std per curve (captures bed boundaries)
    - Neighboring depth values (t-1, t+1) per curve (local continuity)

    Args:
        curves: (n_curves, n_depth) numpy array
        n_depth: number of depth points
        window: rolling window size (default 5)

    Returns:
        X: (n_depth, n_features) numpy array
    """
    n_curves = curves.shape[0]
    half = window // 2
    features = []

    for d in range(n_depth):
        f = []

        for c in range(n_curves):
            f.append(float(curves[c, d]))

        f.append(d / max(n_depth - 1, 1))

        lo = max(0, d - half)
        hi = min(n_depth, d + half + 1)
        for c in range(n_curves):
            f.append(float(np.nanmean(curves[c, lo:hi])))
        for c in range(n_curves):
            f.append(float(np.nanstd(curves[c, lo:hi])))

        for c in range(n_curves):
            f.append(float(curves[c, max(0, d - 1)]))
        for c in range(n_curves):
            f.append(float(curves[c, min(n_depth - 1, d + 1)]))

        features.append(f)

    return np.array(features, dtype=np.float32)